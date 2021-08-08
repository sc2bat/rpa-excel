# 시트별로 관리하는 경우

from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "0000cc" # RGB 형태로 값을 넣어주면 
# Google search 'RGB'
# https://www.w3schools.com/colors/colors_rgb.asp

ws1 = wb.create_sheet("YSheet") # 생성과 동시에 닉변
ws2 = wb.create_sheet("NSheet", 2) # 2번때 idx 값에 sheet 생성

new_ws = wb["NSheet"] # Dict 형태로 sheet 에 접근

print(wb.sheetnames) # 모든 시트 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws) 
target.title = "Copied Sheet"

wb.save("smaple2.xlsx")