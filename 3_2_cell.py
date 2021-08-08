from random import * 
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "dodoSheet"

index = 1 
# 반복문을 이용 랜덤 숫자 채우기
for x in range(1, 11): # 10개 row
    for y in range(1, 11): # 10 개 column
        # ws.cell(column = y, row = x, value = randint(0,100)) # 0 ~ 100 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1

# wb.save("sample4.xlsx")
wb.save("sample3_2.xlsx") # index 