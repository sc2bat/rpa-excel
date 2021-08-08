from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "dodoSheet"

# A1 셀에 1 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없으면 None 출력

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3) , ...
print(ws.cell(row=1, column=1).value) # ws["A1"].value
print(ws.cell(row=1, column=2).value) # ws["B1"].value
print(ws.cell(column=2, row=2).value) # 둘의 서순 관계없음

c = ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10 값을 기입
print(c.value) # ws["C1"].value

wb.save("sample3_1.xlsx")