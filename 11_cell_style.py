# 스타일
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample5_2.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# 높이 너비 조정
ws.column_dimensions["A"].width = 5
ws.row_dimensions[1].height = 50
# wb.save("sample11.xlsx")

# style 적용
# 글색상, 이탤릭, 두겁게
a1.font = Font(color="FF0000", italic=True, bold=True) 
# wb.save("sample11_1.xlsx")
# 글색상 폰트 취소선
b1.font = Font(color="FF0000", name="Arial", strike=True)
# wb.save("sample11_2.xlsx")
# 글색상 글자 크기 밑줄
c1.font = Font(color="0000FF", size=20, underline="single")
# wb.save("sample11_3.xlsx")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top =Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border
# wb.save("sample11_4.xlsx")

# 특정 셀의 색 변경 
for row in ws.rows:
    for cell in row:
        # 각 cell 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if cell.column == 1: # A 열 제외
            continue

        # 셀 정수형 데이터 그리고 80 초과면
        if isinstance(cell.value, int) and cell.value > 80:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경 초록
            cell.font = Font(color="FF0000") # 글 색상

# 틀 고정
ws.freeze_panes = "B2"
wb.save("sample11_5.xlsx")

