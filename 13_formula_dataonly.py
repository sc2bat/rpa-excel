from openpyxl import load_workbook


# wb = load_workbook("sample13.xlsx")
# ws = wb.active

# # 단순한 데이터가 아닌 수식이 나옴
# for row in ws.values:
#     for cell in row:
#         print(cell) 

# 수식이 아닌 실제 데이터 값
# evaluate 계산되지 않은 수식은 None 으로 나옴
wb = load_workbook("sample13.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell) 

