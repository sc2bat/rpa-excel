from openpyxl import Workbook
from random import * # 영어 수학 점수 랜덤 
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 기입
ws.append(["No", "Eng", "Math"])
for i in range(1, 11): # 10개 데이터
    ws.append([i, randint(0, 100), randint(0, 100)])

# 전체 rows
# # print(tuple(ws.rows)) # 1줄씩 튜플로 가져옴
# for row in tuple(ws.rows):
#     # print(row)
#     print(row[1].value)

# 전체 columns
# # print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): # 전체 row
#     print(row[1].value)

# for column in ws.iter_cols():
#     print(column[0].value)

# for row in ws.iter_rows(min_row=1, max_row=5): 
#     print(row[2].value)

# 2번째 ~ 11번째 줄, 2번째 ~ 3번째 열
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): 
    print(row[0].value, row[1].value)


wb.save("sample5_2.xlsx")
