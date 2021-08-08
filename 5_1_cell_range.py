from openpyxl import Workbook
from random import * # 영어 수학 점수 랜덤 
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 기입
ws.append(["No", "Eng", "Math"])
for i in range(1, 11): # 10개 데이터
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_E = ws["B"] # 특성 column 만 가져오기
# print(col_E)
# for cell in col_E:
#     print(cell.value)

# col_range = ws["B:C"] # B, C column 함께 가져오기
# # for cols in col_range:
# #     for cell in cols:
# #         print(cell.value)

# row_title = ws[1] # 1번째 row 만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번째에서 6번째 줄까지
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate) # A/10
        # print(xy, end=" ")
        print(xy[0], end="")
        print(xy[1], end=" ")
    print()

# wb.save("sample5.xlsx")
wb.save("sample5_1.xlsx")
