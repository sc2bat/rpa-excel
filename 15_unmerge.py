from openpyxl import load_workbook
wb = load_workbook("sample14.xlsx")
ws = wb.active

# 병합된 셀을 분리
ws.unmerge_cells("B2:D2")
wb.save("sample15.xlsx")