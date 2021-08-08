from openpyxl import load_workbook
wb = load_workbook("sample5_2.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row =2):
    # 번호 영어 수학
    if int(row[1].value) > 90:
        print(row[0].value, "번 학생은 영어 우수")


for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "Eng":
            cell.value = "Com"

wb.save("sample6_edit.xlsx")