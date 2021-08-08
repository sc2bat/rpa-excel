# 특정 셀 이동

from openpyxl import load_workbook
wb = load_workbook("sample5_2.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 '국어' 영어 수학
# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "Kook" # 비워져있는 셀 항목에 값 부여

# wb.save("sample9.xlsx")

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1)
wb.save("sample9_2.xlsx")